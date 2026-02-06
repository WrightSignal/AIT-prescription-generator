import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
import datetime
from tkcalendar import DateEntry

# --- Allergen Data ---
ALLERGENS = [
    # Mold Group
    {"name": "Aspergillus", "group": "Mold", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},
    {"name": "Alternaria", "group": "Mold", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},
    {"name": "Cladosporium", "group": "Mold", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},
    {"name": "Penicillium", "group": "Mold", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Other"]},

    # Tree Group
    {"name": "Ash", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Birch (Oak)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Cedar", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Hackberry (Elm)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Maple", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Sycamore", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Walnut (Pecan)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Willow (Cottonwood)", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Mulberry", "group": "Tree", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},

    # Grass Group
    {"name": "Timothy", "group": "Grass", "min_volume": 0.1, "max_volume": 0.4, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Johnson", "group": "Grass", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Bermuda", "group": "Grass", "min_volume": 0.3, "max_volume": 1.5, "incompatible_groups": ["Mold", "Other"]},

    # Weed Group
    {"name": "Cocklebur", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Yellow Dock (Sheep Sorrel)", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Kochia (Firebush)", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Lamb's Quarter", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Mugwort", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Pigweed", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "English Plantain", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Russian Thistle", "group": "Weed", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Mold", "Other"]},
    {"name": "Ragweed", "group": "Weed", "min_volume": 0.3, "max_volume": 0.6, "incompatible_groups": ["Mold", "Other"]},

    # Other Group
    {"name": "Cat", "group": "Other", "min_volume": 1.0, "max_volume": 4.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Dog - UF", "group": "Other", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Dog - Epithelium", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Mouse", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Horse", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed", "Mold"]},
    {"name": "Amer. Cockroach", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed"]},
    {"name": "Ger. Cockroach", "group": "Other", "min_volume": 0.5, "max_volume": 1.0, "incompatible_groups": ["Tree", "Grass", "Weed"]},
    {"name": "Dust Mite Mix", "group": "Other", "min_volume": 0.5, "max_volume": 2.0, "incompatible_groups": ["Tree", "Grass", "Weed"]},

     # Venom Group
    {"name": "Honey Bee", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "Yellow Jacket", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "Yellow Faced Hornet", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "White Faced Hornet", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
    {"name": "Wasp", "group": "Venom", "min_volume": 1.0, "max_volume": 1.0, "incompatible_groups": []},
]

class Vial:
    """Represents a single allergy vial."""

    def __init__(self, label):
        self.label = label
        self.allergens = {}  # {allergen_name: volume}
        self.current_volume = 0.0

    def add_allergen(self, allergen_name, volume):
        """Adds an allergen to the vial if compatible and within volume limits.

        Args:
            allergen_name: The name of the allergen.
            volume: The volume of the allergen to add.

        Returns:
            True if the allergen was added successfully, False otherwise.
        """

        allergen_data = next((a for a in ALLERGENS if a["name"] == allergen_name), None)
        if not allergen_data:
            return False  # Allergen not found

        if not self.is_compatible(allergen_data):
            return False  # Incompatible allergen

        if self.current_volume + volume > 5.0:
            return False  # Exceeds volume limit

        if not (allergen_data["min_volume"] <= volume <= allergen_data["max_volume"]):
            return False  # Volume out of range.

        self.allergens[allergen_name] = volume
        self.current_volume += volume
        return True

    def remaining_volume(self):
        """Calculates the remaining volume in the vial."""
        return 5.0 - self.current_volume

    def is_compatible(self, allergen_data):
        """Checks if an allergen is compatible with the current vial contents."""
        current_groups = {a["group"] for a in ALLERGENS if a["name"] in self.allergens}
        for group in current_groups:
            if group in allergen_data["incompatible_groups"]:
                return False
        if allergen_data["group"] in {a["group"] for a in ALLERGENS for current_a in self.allergens if a["name"] == current_a}:
          return False
        return True

    def get_contents_string(self):
        """Returns a formatted string of the vial's contents."""
        contents = []
        for allergen, volume in self.allergens.items():
            contents.append(f"  - {allergen}: {volume:.2f} mL")
        contents.append(f"  - Diluent: {self.remaining_volume():.2f} mL")
        return "\n".join(contents)


def save_patient_data(patient_data, selected_allergens):
    """Saves patient data and selected allergens to the Excel file."""
    try:
        try:
            workbook = load_workbook(filename="patient_data.xlsx")
            sheet = workbook["Sheet1"]
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["Patient Name", "Date of Birth", "MRN", "Street Address", "City", "State", "Phone Number", "Allergens"])

        # Check for existing patient (using MRN)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == patient_data['mrn']:
                response = messagebox.askyesno("Patient Exists",
                                             f"Patient with MRN '{patient_data['mrn']}' already exists. Overwrite?")
                if response:
                    row_num_to_delete = 0
                    for i,r in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if r[2] == patient_data['mrn']:
                            row_num_to_delete = i
                            break
                    if row_num_to_delete:
                        sheet.delete_rows(row_num_to_delete)
                else:
                    return

        # Append/Update new patient data, including allergens
        sheet.append([
            patient_data['patient_name'],
            patient_data['dob'].strftime("%m-%d-%Y"),
            patient_data['mrn'],
            patient_data['address'],
            patient_data['city'],
            patient_data['state'],
            patient_data['phone'],
            selected_allergens  # Comma-separated string of allergens
        ])
        workbook.save(filename="patient_data.xlsx")
        messagebox.showinfo("Success", f"Patient '{patient_data['patient_name']}' saved successfully.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving: {e}")

def load_patient():
    """Opens a new window to load an existing patient."""

    def load_selected_patient():
        selected_patient_str = patient_select_var.get()
        if not selected_patient_str:
            return

        try:
            # Split the string into name and DOB parts
            name_part, dob_part = selected_patient_str.rsplit(" ", 1)  # Split on the last space
            dob_part = dob_part.strip() #Remove extra space if there is any.
            name_part = name_part.strip()

            # Convert the DOB string to a date object for comparison
            dob_to_match = datetime.datetime.strptime(dob_part, "%m-%d-%Y").date()


            workbook = load_workbook(filename="patient_data.xlsx", read_only=True)
            sheet = workbook["Sheet1"]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Compare both name and DOB
                try:
                    row_dob = datetime.datetime.strptime(row[1], "%m-%d-%Y").date()
                except ValueError:
                    #Handle possible date conversion issues (e.g., bad data in Excel)
                    continue  # Skip this row if the date is invalid

                if row[0] == name_part and row_dob == dob_to_match:
                    # Populate the main window's fields
                    patient_name_entry.delete(0, tk.END)
                    patient_name_entry.insert(0, row[0])

                    # Date of Birth
                    dob_date = datetime.datetime.strptime(row[1], "%m-%d-%Y").date()
                    dob_entry.set_date(dob_date)

                    mrn_entry.delete(0, tk.END)
                    mrn_entry.insert(0, row[2])
                    address_entry.delete(0, tk.END)
                    address_entry.insert(0, row[3])
                    city_entry.delete(0, tk.END)
                    city_entry.insert(0, row[4])
                    state_entry.delete(0, tk.END)
                    state_entry.insert(0, row[5])
                    phone_entry.delete(0, tk.END)
                    phone_entry.insert(0, row[6])

                    # Load and set allergen checkboxes
                    allergens_str = row[7]
                    selected_allergens = allergens_str.split(",") if allergens_str else []

                    # Clear current checkbox selections
                    for var in environmental_allergen_vars.values():
                        var.set(False)
                    for var in venom_allergen_vars.values():
                        var.set(False)

                    #Set checkboxes based on loaded data:
                    if selected_allergens is not None:
                        for allergen in selected_allergens:
                            allergen = allergen.strip()
                            if allergen in environmental_allergen_vars:
                                environmental_allergen_vars[allergen].set(True)
                            elif allergen in venom_allergen_vars:
                                venom_allergen_vars[allergen].set(True)

                    # Set the Vial Type based on which allergens are selected
                    if any(allergen in venom_allergens for allergen in selected_allergens):
                        vial_type_var.set("Venom")
                    elif any(allergen in environmental_allergens for allergen in selected_allergens):
                        vial_type_var.set("Environmental")
                    else:
                        vial_type_var.set("Environmental")

                    load_window.destroy()  # Close the load window
                    return

            messagebox.showwarning("Patient Not Found", f"No patient found with Name and DOB: {selected_patient_str}")

        except FileNotFoundError:
            messagebox.showerror("Error", "patient_data.xlsx not found!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    # --- Create the Load Patient Window ---
    load_window = tk.Toplevel(root)
    load_window.title("Load Patient")

    # Patient Selection Dropdown
    patient_select_label = ttk.Label(load_window, text="Select Patient (Name DOB):")
    patient_select_label.pack(padx=10, pady=5)

    patient_select_var = tk.StringVar()
    patient_select_combo = ttk.Combobox(load_window, textvariable=patient_select_var)
    patient_select_combo.pack(padx=10, pady=5)

    # Populate the dropdown with existing patient names and DOBs
    try:
        workbook = load_workbook(filename="patient_data.xlsx", read_only=True)
        sheet = workbook["Sheet1"]
        patient_strings = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                name = row[0]
                dob_str = row[1]
                # Format the DOB string for the dropdown
                dob_formatted = datetime.datetime.strptime(dob_str, "%m-%d-%Y").strftime("%m-%d-%Y")
                patient_strings.append(f"{name} {dob_formatted}")
            except ValueError:
                # Handle cases where the date in Excel might be invalid
                continue
        patient_select_combo['values'] = patient_strings

    except FileNotFoundError:
        messagebox.showinfo("No Patients", "No patient data found.  Add patients first.")
        load_window.destroy()
        return
    except Exception as e:
        messagebox.showerror("Error", f"Error loading patient list: {e}")
        load_window.destroy()
        return

    # Load Button
    load_button = ttk.Button(load_window, text="Load", command=load_selected_patient)
    load_button.pack(pady=10)


def calculate_vials(selected_allergens):
    """Calculates vial assignments based on allergen compatibility and volumes.
    
    Returns a list of Vial objects with allergens assigned.
    """
    if not selected_allergens:
        return []
    
    # Get allergen data for selected allergens
    allergen_data_list = []
    for allergen_name in selected_allergens:
        allergen_data = next((a for a in ALLERGENS if a["name"] == allergen_name), None)
        if allergen_data:
            allergen_data_list.append(allergen_data)
    
    # Group allergens by compatibility
    # Mold must be separate from Tree, Grass, Weed, Other
    # Other (animals/insects) must be separate from Tree, Grass, Weed, Mold
    # Tree, Grass, Weed can be mixed together
    # Venom is always separate (each venom in its own vial typically)
    
    groups = {
        "Mold": [],
        "Pollen": [],  # Tree, Grass, Weed combined
        "Other": [],
        "Venom": []
    }
    
    for allergen_data in allergen_data_list:
        group = allergen_data["group"]
        if group == "Mold":
            groups["Mold"].append(allergen_data)
        elif group in ["Tree", "Grass", "Weed"]:
            groups["Pollen"].append(allergen_data)
        elif group == "Other":
            groups["Other"].append(allergen_data)
        elif group == "Venom":
            groups["Venom"].append(allergen_data)
    
    vials = []
    vial_number = 1
    
    # Process each compatibility group
    for group_name, allergens in groups.items():
        if not allergens:
            continue
        
        # For venom, each allergen gets its own vial
        if group_name == "Venom":
            for allergen_data in allergens:
                vial = Vial(f"Vial {vial_number} ({group_name})")
                vial.add_allergen(allergen_data["name"], allergen_data["min_volume"])
                vials.append(vial)
                vial_number += 1
        else:
            # For other groups, pack allergens into vials up to 5mL
            current_vial = Vial(f"Vial {vial_number} ({group_name})")
            
            for allergen_data in allergens:
                volume = allergen_data["min_volume"]
                
                # Check if allergen fits in current vial
                if current_vial.current_volume + volume <= 5.0:
                    current_vial.allergens[allergen_data["name"]] = volume
                    current_vial.current_volume += volume
                else:
                    # Save current vial and start a new one
                    if current_vial.allergens:
                        vials.append(current_vial)
                        vial_number += 1
                    current_vial = Vial(f"Vial {vial_number} ({group_name})")
                    current_vial.allergens[allergen_data["name"]] = volume
                    current_vial.current_volume += volume
            
            # Don't forget to add the last vial
            if current_vial.allergens:
                vials.append(current_vial)
                vial_number += 1
    
    return vials


def generate_prescription():
    """Generates the prescription text."""
    mode = vial_type_var.get()
    patient_name = patient_name_entry.get()
    mrn = mrn_entry.get()
    address = address_entry.get()
    city = city_entry.get()
    state = state_entry.get()
    phone = phone_entry.get()

    if not patient_name or not mrn:
        messagebox.showerror("Error", "Please enter patient name and MRN.")
        return

    try:
        dob = dob_entry.get_date()
    except ValueError:
        messagebox.showerror("Error", "Invalid date of birth entered.")
        return

    patient_data = {
        'patient_name': patient_name,
        'dob': dob,
        'mrn': mrn,
        'address': address,
        'city': city,
        'state': state,
        'phone': phone
    }

    # --- Collect Selected Allergens ---
    if mode == "Environmental":
        selected_allergens = [
            allergen for allergen, var in environmental_allergen_vars.items() if var.get()
        ]
    elif mode == "Venom":
        selected_allergens = [
            allergen for allergen, var in venom_allergen_vars.items() if var.get()
        ]
    else:
        result_label.config(text="Invalid vial type selected.")
        return
    # Convert to comma-separated string
    selected_allergens_str = ", ".join(selected_allergens)

    save_patient_data(patient_data, selected_allergens_str)  # Pass the allergens

    prescription_text = f"Patient Name: {patient_name}\n"
    prescription_text += f"Date of Birth: {dob.strftime('%m-%d-%Y')}\n"
    prescription_text += f"MRN: {mrn}\n"
    prescription_text += f"Address: {address}\n"
    prescription_text += f"City: {city}, {state}\n"
    prescription_text += f"Phone: {phone}\n"
    prescription_text += f"Vial Type: {mode}\n"
    prescription_text += "-" * 40 + "\n"

    if not selected_allergens:
        prescription_text += "  (No allergens selected)\n"
    else:
        # Calculate vial assignments
        vials = calculate_vials(selected_allergens)
        
        prescription_text += f"VIAL FORMULATIONS ({len(vials)} vial(s) total):\n"
        prescription_text += "-" * 40 + "\n"
        
        for vial in vials:
            prescription_text += f"\n{vial.label}:\n"
            prescription_text += "  Stock Extracts:\n"
            
            total_extract = 0.0
            for allergen_name, volume in vial.allergens.items():
                prescription_text += f"    • {allergen_name}: {volume:.2f} mL\n"
                total_extract += volume
            
            diluent_volume = vial.remaining_volume()
            prescription_text += f"  ─────────────────────\n"
            prescription_text += f"  Total Stock Extract: {total_extract:.2f} mL\n"
            prescription_text += f"  Diluent (HSA): {diluent_volume:.2f} mL\n"
            prescription_text += f"  Final Vial Volume: 5.00 mL\n"

    result_label.config(text=prescription_text)


def update_allergen_options(*args):
    """Updates the visible allergen checkboxes based on vial type."""
    mode = vial_type_var.get()

    if mode == "Environmental":
        for widget in venom_allergen_frame.winfo_children():
            widget.grid_remove()
        for frame in environmental_allergen_frames:
            frame.grid()

    elif mode == "Venom":
        for frame in environmental_allergen_frames:
            frame.grid_remove()
        for i, allergen in enumerate(venom_allergens):
            venom_allergen_checkboxes[i].grid()

    else:
        for frame in environmental_allergen_frames:
            frame.grid_remove()
        for widget in venom_allergen_frame.winfo_children():
            widget.grid_remove()


def clear_fields():
    """Clears all input fields and checkbox selections."""
    patient_name_entry.delete(0, tk.END)
    dob_entry.set_date(datetime.date.today())
    mrn_entry.delete(0, tk.END)
    address_entry.delete(0, tk.END)
    city_entry.delete(0, tk.END)
    state_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    # Uncheck all checkboxes
    for var in environmental_allergen_vars.values():
        var.set(False)
    for var in venom_allergen_vars.values():
        var.set(False)
    result_label.config(text="")  # Clear result label


# --- Main Application Window ---
root = tk.Tk()
root.title("AIT Prescription Generator")
root.geometry("900x700")
root.configure(bg="#f0f4f8")

# --- Custom Styling ---
style = ttk.Style()
style.theme_use('clam')

# Define colors
PRIMARY_COLOR = "#2563eb"  # Blue
PRIMARY_HOVER = "#1d4ed8"
SECONDARY_COLOR = "#64748b"  # Slate
BG_COLOR = "#f8fafc"
CARD_BG = "#ffffff"
TEXT_COLOR = "#1e293b"
ACCENT_COLOR = "#10b981"  # Green for success

# Configure styles
style.configure("TFrame", background=BG_COLOR)
style.configure("Card.TFrame", background=CARD_BG, relief="flat")
style.configure("TLabelframe", background=CARD_BG, foreground=TEXT_COLOR, borderwidth=2, relief="groove")
style.configure("TLabelframe.Label", background=CARD_BG, foreground=PRIMARY_COLOR, font=("Segoe UI", 10, "bold"))
style.configure("TLabel", background=CARD_BG, foreground=TEXT_COLOR, font=("Segoe UI", 10))
style.configure("Header.TLabel", background=BG_COLOR, foreground=TEXT_COLOR, font=("Segoe UI", 16, "bold"))
style.configure("TEntry", font=("Segoe UI", 10), padding=5)
style.configure("TCheckbutton", background=CARD_BG, foreground=TEXT_COLOR, font=("Segoe UI", 9))
style.configure("TCombobox", font=("Segoe UI", 10), padding=5)

# Button styles
style.configure("Primary.TButton", 
                background=PRIMARY_COLOR, 
                foreground="white", 
                font=("Segoe UI", 10, "bold"),
                padding=(15, 8))
style.map("Primary.TButton",
          background=[("active", PRIMARY_HOVER), ("pressed", PRIMARY_HOVER)])

style.configure("Secondary.TButton",
                background=SECONDARY_COLOR,
                foreground="white",
                font=("Segoe UI", 10),
                padding=(15, 8))
style.map("Secondary.TButton",
          background=[("active", "#475569"), ("pressed", "#475569")])

style.configure("Accent.TButton",
                background=ACCENT_COLOR,
                foreground="white",
                font=("Segoe UI", 10, "bold"),
                padding=(15, 8))

# --- Main Container with Scrollbar ---
main_frame = ttk.Frame(root, style="TFrame")
main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

my_canvas = tk.Canvas(main_frame, bg=BG_COLOR, highlightthickness=0)
my_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

my_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=my_canvas.yview)
my_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

my_canvas.configure(yscrollcommand=my_scrollbar.set)
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))

# Mouse wheel scrolling
def _on_mousewheel(event):
    my_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
my_canvas.bind_all("<MouseWheel>", _on_mousewheel)

# --- Content Frame ---
second_frame = ttk.Frame(my_canvas, style="TFrame")
my_canvas.create_window((0, 0), window=second_frame, anchor="nw")

# --- Header ---
header_frame = ttk.Frame(second_frame, style="TFrame")
header_frame.grid(row=0, column=0, columnspan=2, pady=(0, 15), sticky="ew")

header_label = ttk.Label(header_frame, text="🩺 AIT Prescription Generator", style="Header.TLabel")
header_label.pack(side=tk.LEFT)

# --- Two Column Layout ---
left_column = ttk.Frame(second_frame, style="TFrame")
left_column.grid(row=1, column=0, padx=(0, 10), sticky="nsew")

right_column = ttk.Frame(second_frame, style="TFrame")
right_column.grid(row=1, column=1, padx=(10, 0), sticky="nsew")

# --- Patient Information Card ---
patient_info_frame = ttk.LabelFrame(left_column, text="  Patient Information  ", padding=(15, 10))
patient_info_frame.grid(row=0, column=0, pady=(0, 10), sticky="ew")

# Configure grid weights for patient info
patient_info_frame.columnconfigure(1, weight=1)

# Patient Name
patient_name_label = ttk.Label(patient_info_frame, text="Patient Name:")
patient_name_label.grid(row=0, column=0, padx=(0, 10), pady=8, sticky="w")
patient_name_entry = ttk.Entry(patient_info_frame, width=30)
patient_name_entry.grid(row=0, column=1, pady=8, sticky="ew")

# Date of Birth
dob_label = ttk.Label(patient_info_frame, text="Date of Birth:")
dob_label.grid(row=1, column=0, padx=(0, 10), pady=8, sticky="w")
dob_entry = DateEntry(patient_info_frame, width=18, background=PRIMARY_COLOR, foreground='white',
                      borderwidth=2, date_pattern='m-d-Y', font=("Segoe UI", 10))
dob_entry.grid(row=1, column=1, pady=8, sticky="w")

# Medical Record Number
mrn_label = ttk.Label(patient_info_frame, text="MRN:")
mrn_label.grid(row=2, column=0, padx=(0, 10), pady=8, sticky="w")
mrn_entry = ttk.Entry(patient_info_frame, width=30)
mrn_entry.grid(row=2, column=1, pady=8, sticky="ew")

# Street Address
address_label = ttk.Label(patient_info_frame, text="Street Address:")
address_label.grid(row=3, column=0, padx=(0, 10), pady=8, sticky="w")
address_entry = ttk.Entry(patient_info_frame, width=30)
address_entry.grid(row=3, column=1, pady=8, sticky="ew")

# City and State in one row
city_state_frame = ttk.Frame(patient_info_frame)
city_state_frame.grid(row=4, column=0, columnspan=2, pady=8, sticky="ew")

city_label = ttk.Label(city_state_frame, text="City:")
city_label.grid(row=0, column=0, padx=(0, 10), sticky="w")
city_entry = ttk.Entry(city_state_frame, width=20)
city_entry.grid(row=0, column=1, padx=(0, 15), sticky="w")

state_label = ttk.Label(city_state_frame, text="State:")
state_label.grid(row=0, column=2, padx=(0, 10), sticky="w")
state_entry = ttk.Entry(city_state_frame, width=8)
state_entry.grid(row=0, column=3, sticky="w")

# Phone Number
phone_label = ttk.Label(patient_info_frame, text="Phone Number:")
phone_label.grid(row=5, column=0, padx=(0, 10), pady=8, sticky="w")
phone_entry = ttk.Entry(patient_info_frame, width=30)
phone_entry.grid(row=5, column=1, pady=8, sticky="ew")

# --- Vial Type Selection ---
vial_selection_frame = ttk.LabelFrame(left_column, text="  Vial Configuration  ", padding=(15, 10))
vial_selection_frame.grid(row=1, column=0, pady=10, sticky="ew")

vial_type_label = ttk.Label(vial_selection_frame, text="Vial Type:")
vial_type_label.grid(row=0, column=0, padx=(0, 10), pady=8, sticky="w")
vial_type_var = tk.StringVar(value="Environmental")
vial_type_var.trace_add("write", update_allergen_options)

vial_type_combo = ttk.Combobox(vial_selection_frame, textvariable=vial_type_var,
                               values=["Environmental", "Venom"], width=18, state="readonly")
vial_type_combo.grid(row=0, column=1, pady=8, sticky="w")

load_patient_button = ttk.Button(vial_selection_frame, text="📂 Load Patient", 
                                  command=load_patient, style="Secondary.TButton")
load_patient_button.grid(row=0, column=2, padx=(20, 0), pady=8, sticky="e")

# --- Allergen Selection Area ---
allergen_container = ttk.LabelFrame(left_column, text="  Allergen Selection  ", padding=(15, 10))
allergen_container.grid(row=2, column=0, pady=10, sticky="ew")

# Environmental Allergen Checkboxes
environmental_allergen_frames = []
environmental_allergen_vars = {}

# --- Mold Group ---
mold_frame = ttk.LabelFrame(allergen_container, text="  Mold  ", padding=(10, 5))
mold_frame.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
environmental_allergen_frames.append(mold_frame)
mold_allergens = ["Aspergillus", "Alternaria", "Cladosporium", "Penicillium"]
for i, allergen in enumerate(mold_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(mold_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 2, column=i % 2, padx=10, pady=4, sticky="w")

# --- Tree Group ---
tree_frame = ttk.LabelFrame(allergen_container, text="  Tree  ", padding=(10, 5))
tree_frame.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
environmental_allergen_frames.append(tree_frame)
tree_allergens = ["Ash", "Birch (Oak)", "Cedar", "Hackberry (Elm)", "Maple", "Sycamore", "Walnut (Pecan)",
                  "Willow (Cottonwood)", "Mulberry"]
for i, allergen in enumerate(tree_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(tree_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 3, column=i % 3, padx=10, pady=4, sticky="w")

# --- Grass Group ---
grass_frame = ttk.LabelFrame(allergen_container, text="  Grass  ", padding=(10, 5))
grass_frame.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
environmental_allergen_frames.append(grass_frame)
grass_allergens = ["Timothy", "Johnson", "Bermuda"]
for i, allergen in enumerate(grass_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(grass_frame, text=allergen, variable=var)
    checkbox.grid(row=0, column=i, padx=10, pady=4, sticky="w")

# --- Weed Group ---
weed_frame = ttk.LabelFrame(allergen_container, text="  Weed  ", padding=(10, 5))
weed_frame.grid(row=3, column=0, padx=5, pady=5, sticky="ew")
environmental_allergen_frames.append(weed_frame)
weed_allergens = ["Cocklebur", "Yellow Dock (Sheep Sorrel)", "Kochia (Firebush)", "Lamb's Quarter", "Mugwort",
                  "Pigweed", "English Plantain", "Russian Thistle", "Ragweed"]
for i, allergen in enumerate(weed_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(weed_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 3, column=i % 3, padx=10, pady=4, sticky="w")

# --- Other Group ---
other_frame = ttk.LabelFrame(allergen_container, text="  Other (Animals/Insects)  ", padding=(10, 5))
other_frame.grid(row=4, column=0, padx=5, pady=5, sticky="ew")
environmental_allergen_frames.append(other_frame)
other_allergens = ["Cat", "Dog - UF", "Dog - Epithelium", "Mouse", "Horse", "Amer. Cockroach", "Ger. Cockroach",
                   "Dust Mite Mix"]
for i, allergen in enumerate(other_allergens):
    var = tk.BooleanVar()
    environmental_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(other_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 3, column=i % 3, padx=10, pady=4, sticky="w")

# --- Venom Allergen Checkboxes ---
venom_allergen_frame = ttk.LabelFrame(allergen_container, text="  Venom Allergens  ", padding=(10, 5))
venom_allergen_frame.grid(row=5, column=0, padx=5, pady=5, sticky="ew")

venom_allergens = ["Honey Bee", "Yellow Jacket", "Yellow Faced Hornet", "White Faced Hornet", "Wasp"]
venom_allergen_vars = {}
venom_allergen_checkboxes = []

for i, allergen in enumerate(venom_allergens):
    var = tk.BooleanVar()
    venom_allergen_vars[allergen] = var
    checkbox = ttk.Checkbutton(venom_allergen_frame, text=allergen, variable=var)
    checkbox.grid(row=i // 3, column=i % 3, padx=10, pady=4, sticky="w")
    venom_allergen_checkboxes.append(checkbox)

# Initially hide venom allergens
for widget in venom_allergen_frame.winfo_children():
    widget.grid_remove()

# --- Action Buttons ---
button_frame = ttk.Frame(left_column, style="TFrame")
button_frame.grid(row=3, column=0, pady=15, sticky="ew")

generate_button = ttk.Button(button_frame, text="✨ Generate Prescription", 
                             command=generate_prescription, style="Primary.TButton")
generate_button.pack(side=tk.LEFT, padx=(0, 10))

clear_button = ttk.Button(button_frame, text="🗑️ Clear Fields", 
                          command=clear_fields, style="Secondary.TButton")
clear_button.pack(side=tk.LEFT)

# --- Right Column: Prescription Output ---
output_frame = ttk.LabelFrame(right_column, text="  Prescription Output  ", padding=(15, 10))
output_frame.grid(row=0, column=0, sticky="nsew")

# Use a Text widget with better formatting for the output
result_text = tk.Text(output_frame, wrap=tk.WORD, width=45, height=35, 
                      font=("Consolas", 10), bg="#f8fafc", fg=TEXT_COLOR,
                      relief="flat", padx=10, pady=10)
result_text.pack(fill=tk.BOTH, expand=True)
result_text.config(state=tk.DISABLED)

# Create a wrapper to update the text widget instead of label
def update_result_display(text):
    result_text.config(state=tk.NORMAL)
    result_text.delete(1.0, tk.END)
    result_text.insert(tk.END, text)
    result_text.config(state=tk.DISABLED)

# Replace result_label with a dummy for compatibility
class ResultLabelWrapper:
    def config(self, text=""):
        update_result_display(text)

result_label = ResultLabelWrapper()

# Configure grid weights
second_frame.columnconfigure(0, weight=1)
second_frame.columnconfigure(1, weight=1)
right_column.rowconfigure(0, weight=1)

root.mainloop()
